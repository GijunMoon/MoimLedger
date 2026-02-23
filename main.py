### MoimLedger
## Author: Gijun Moon
## Date: 26-02-23
### main.py

from fastapi import FastAPI, Request, UploadFile, Form, File
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
import pandas as pd
import io
import os
from openpyxl import Workbook
from datetime import datetime
from collections import defaultdict

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import TableStyle

import ocr_module as ocr

import uuid

temporary_storage = {}

def _format_krw(value: float) -> str:
    if float(value).is_integer():
        return f"{int(value):,}원"
    return f"{value:,.2f}원"


def _signed_krw(value: float) -> str:
    if value > 0:
        return f"+{_format_krw(value)}"
    if value < 0:
        return f"-{_format_krw(abs(value))}"
    return _format_krw(0)


def _resolve_korean_pdf_font() -> str:
    font_candidates = [
        ("MalgunGothic", r"C:\Windows\Fonts\malgun.ttf"),
        ("AppleGothic", "/System/Library/Fonts/Supplemental/AppleGothic.ttf"),
        ("NotoSansCJKkr", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
    ]

    for font_name, font_path in font_candidates:
        if os.path.exists(font_path):
            try:
                if font_name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                return font_name
            except Exception:
                continue

    cid_font_name = "HYSMyeongJo-Medium"
    if cid_font_name not in pdfmetrics.getRegisteredFontNames():
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(cid_font_name))
        except Exception:
            return "Helvetica"

    return cid_font_name

app = FastAPI()
templates = Jinja2Templates(directory="templates")

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse(
        "main.html",
        {
            "request": request
        }
    )

@app.post("/create-session", response_class=HTMLResponse)
async def create_session(request: Request, members: str = Form(...)):

    member_list = [m.strip() for m in members.split(",") if m.strip()]

    session_id = str(uuid.uuid4())

    temporary_storage[session_id] = {
        "members": member_list,
        "transactions": [],
        "settlement": [],
        "transfers": [],
        "total_amount": 0,
        "per_person": 0
    }

    recalculate(session_id)
    return RedirectResponse(url=f"/upload/{session_id}", status_code=303)


@app.get("/upload/{session_id}", response_class=HTMLResponse)
def upload_page(request: Request, session_id: str):
    if session_id not in temporary_storage:
        return RedirectResponse(url="/", status_code=303)

    result = recalculate(session_id)
    return templates.TemplateResponse(
        "upload.html",
        {
            "request": request,
            "result": result,
            "error": None
        }
    )

@app.post("/upload-csv/{session_id}", response_class=HTMLResponse)
async def upload_csv(
    request: Request,
    session_id: str,
    file: UploadFile
):

    if session_id not in temporary_storage:
        return RedirectResponse(url="/", status_code=303)

    contents = await file.read()
    df = pd.read_csv(io.BytesIO(contents))

    if "amount" not in df.columns or "payer" not in df.columns:
        result = recalculate(session_id)
        return templates.TemplateResponse(
            "upload.html",
            {
                "request": request,
                "result": result,
                "error": "amount, payer 컬럼이 필요합니다."
            }
        )

    transactions = df.to_dict("records")
    temporary_storage[session_id]["transactions"].extend(transactions)

    recalculate(session_id)
    return RedirectResponse(url=f"/upload/{session_id}", status_code=303)


@app.post("/ocr/{session_id}", response_class=HTMLResponse)
async def ocr_receipt(
    request: Request,
    session_id: str,
    payer: str = Form(...),
    file: UploadFile = File(...)
):

    if session_id not in temporary_storage:
        return RedirectResponse(url="/", status_code=303)

    contents = await file.read()
    result = ocr.read(contents)

    for t in result["raw_data"]:
        t["payer"] = payer

    temporary_storage[session_id]["transactions"].extend(result["raw_data"])

    recalculate(session_id)
    return RedirectResponse(url=f"/upload/{session_id}", status_code=303)


@app.get("/result/{session_id}", response_class=HTMLResponse)
def result_page(request: Request, session_id: str):
    if session_id not in temporary_storage:
        return RedirectResponse(url="/", status_code=303)

    result = recalculate(session_id)
    return templates.TemplateResponse(
        "result.html",
        {
            "request": request,
            "result": result
        }
    )
@app.get("/download-excel/{session_id}")
def download_excel(session_id: str):
    data = temporary_storage.get(session_id)
    if not data:
        return {"error": "session not found"}

    settlement = data["settlement"]
    transfers = data["transfers"]

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "summary"
    ws_summary.append(["total_amount", data["total_amount"]])
    ws_summary.append(["per_person", data["per_person"]])

    ws_settlement = wb.create_sheet("settlement")
    ws_settlement.append(["name", "paid", "balance"])
    for s in settlement:
        ws_settlement.append([s["name"], s["paid"], s["balance"]])

    ws_transfer = wb.create_sheet("transfers")
    ws_transfer.append(["from", "to", "amount"])
    for t in transfers:
        ws_transfer.append([t["from"], t["to"], t["amount"]])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=moim_settlement.xlsx"}
    )


@app.get("/download-pdf/{session_id}")
def download_pdf(session_id: str):
    data = temporary_storage.get(session_id)
    if not data:
        return {"error": "session not found"}

    font_name = _resolve_korean_pdf_font()
    settlement = data["settlement"]
    transfers = data["transfers"]
    total_amount = data["total_amount"]
    per_person = data["per_person"]
    members = data.get("members", [])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=40,
        rightMargin=40,
        topMargin=46,
        bottomMargin=44,
        title="모임 정산 장부",
    )
    elements = []

    usable_width = A4[0] - doc.leftMargin - doc.rightMargin

    title_style = ParagraphStyle(
        name="DocTitle",
        fontName=font_name,
        fontSize=20,
        leading=24,
        alignment=1,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        name="DocSubtitle",
        fontName=font_name,
        fontSize=10,
        leading=14,
        alignment=1,
        textColor=colors.HexColor("#3F3F46"),
        spaceAfter=10,
    )
    section_style = ParagraphStyle(
        name="Section",
        fontName=font_name,
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#111827"),
        spaceBefore=8,
        spaceAfter=4,
    )
    note_style = ParagraphStyle(
        name="Note",
        fontName=font_name,
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#4B5563"),
    )

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    document_no = f"ML-{session_id[:8].upper()}"

    elements.append(Paragraph("모임 정산 장부", title_style))
    elements.append(Paragraph("공식 정산 보고서", subtitle_style))

    meta_table = Table(
        [
            ["문서번호", document_no, "작성일시", generated_at],
            ["세션 ID", session_id, "참여인원", f"{len(members)}명"],
        ],
        colWidths=[60, usable_width * 0.38, 60, usable_width * 0.38],
    )
    meta_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F3F4F6")),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#9CA3AF")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(meta_table)
    elements.append(Spacer(1, 0.22 * inch))

    elements.append(Paragraph("1. 정산 요약", section_style))
    summary_table = Table(
        [
            ["총 지출액", _format_krw(total_amount)],
            ["1인당 부담액", _format_krw(per_person)],
            ["정산 대상 인원", f"{len(settlement)}명"],
            ["송금 권장 건수", f"{len(transfers)}건"],
        ],
        colWidths=[110, usable_width - 110],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EEF2FF")),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#94A3B8")),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 0.22 * inch))

    elements.append(Paragraph("2. 개인별 정산 내역", section_style))
    settlement_rows = [["번호", "이름", "실제 결제액", "정산 차액", "상태"]]
    for idx, s in enumerate(settlement, start=1):
        balance = s["balance"]
        if balance > 0:
            status = "받을 금액"
        elif balance < 0:
            status = "보낼 금액"
        else:
            status = "정산 완료"

        settlement_rows.append(
            [
                str(idx),
                s["name"],
                _format_krw(s["paid"]),
                _signed_krw(balance),
                status,
            ]
        )

    settlement_table = Table(
        settlement_rows,
        colWidths=[42, 90, 120, 120, usable_width - (42 + 90 + 120 + 120)],
        repeatRows=1,
    )
    settlement_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9CA3AF")),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 1), (3, -1), "RIGHT"),
                ("ALIGN", (4, 1), (4, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.append(settlement_table)
    elements.append(Spacer(1, 0.22 * inch))

    elements.append(Paragraph("3. 권장 송금 내역", section_style))
    if transfers:
        transfer_rows = [["송금인", "수금인", "송금액"]]
        for t in transfers:
            transfer_rows.append([t["from"], t["to"], _format_krw(t["amount"])])

        transfer_table = Table(
            transfer_rows,
            colWidths=[usable_width * 0.34, usable_width * 0.34, usable_width * 0.32],
            repeatRows=1,
        )
        transfer_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9CA3AF")),
                    ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(transfer_table)
    else:
        no_transfer_table = Table(
            [["추가 송금이 필요하지 않습니다."]],
            colWidths=[usable_width],
        )
        no_transfer_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9CA3AF")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements.append(no_transfer_table)

    elements.append(Spacer(1, 0.28 * inch))
    elements.append(
        Paragraph(
            "본 문서는 Moim Ledger 시스템에서 자동 생성된 문서이며, 입력된 거래 내역을 기준으로 산출되었습니다.",
            note_style,
        )
    )

    def _draw_footer(canvas, _doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#9CA3AF"))
        canvas.setLineWidth(0.5)
        canvas.line(doc.leftMargin, 30, A4[0] - doc.rightMargin, 30)
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor("#6B7280"))
        canvas.drawString(doc.leftMargin, 18, f"Moim Ledger | 문서번호 {document_no}")
        canvas.drawRightString(A4[0] - doc.rightMargin, 18, f"페이지 {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=moim_settlement.pdf"}
    )

def calculate_min_transfers(settlement):
    creditors = []
    debtors = []

    for s in settlement:
        if s["balance"] > 0:
            creditors.append({"name": s["name"], "amount": s["balance"]})
        elif s["balance"] < 0:
            debtors.append({"name": s["name"], "amount": -s["balance"]})

    transfers = []

    i, j = 0, 0

    while i < len(debtors) and j < len(creditors):
        debtor = debtors[i]
        creditor = creditors[j]

        transfer_amount = min(debtor["amount"], creditor["amount"])

        transfers.append({
            "from": debtor["name"],
            "to": creditor["name"],
            "amount": round(transfer_amount, 2)
        })

        debtor["amount"] -= transfer_amount
        creditor["amount"] -= transfer_amount

        if debtor["amount"] == 0:
            i += 1
        if creditor["amount"] == 0:
            j += 1

    return transfers

def recalculate(session_id):
    data = temporary_storage[session_id]

    transactions = data["transactions"]
    members = data["members"]

    total_amount = sum(t["amount"] for t in transactions)
    per_person = total_amount / len(members) if members else 0

    paid_dict = defaultdict(float)

    for t in transactions:
        paid_dict[t["payer"]] += t["amount"]

    settlement = []

    for member in members:
        paid = paid_dict.get(member, 0)
        balance = round(paid - per_person, 2)

        settlement.append({
            "name": member,
            "paid": paid,
            "balance": balance
        })

    transfers = calculate_min_transfers(settlement)

    data["df"] = pd.DataFrame(transactions)
    data["settlement"] = settlement
    data["transfers"] = transfers
    data["total_amount"] = total_amount
    data["per_person"] = per_person

    return {
        "session_id": session_id,
        "total_amount": round(total_amount, 2),
        "per_person": round(per_person, 2),
        "settlement": settlement,
        "transfers": transfers,
        "member_list": members
    }




